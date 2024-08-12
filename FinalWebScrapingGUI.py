from tkinter import *
from PIL import ImageTk, Image
import os
import pathlib
import openpyxl, xlrd
from openpyxl import Workbook
import subprocess
from tkinter import messagebox
import requests,openpyxl
from bs4 import BeautifulSoup
from newspaper import Article
from datetime import date

os.chdir("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT")
def usadmin():
    usad = Toplevel(root)
    usad.title("Web Scraping using Python")
    usad.geometry("1280x720")

    #create two frames
    frame1 = Frame(usad)
    frame1.grid(row=0, column=0, sticky='nsew')
    frame2 = Frame(usad)
    frame2.grid(row=0, column=1, sticky='nsew')

    # Create two canvases, one in each frame
    canvas1 = Canvas(frame1, bg='dark turquoise', width=640, height=720)
    canvas1.pack()
    img = PhotoImage(file='admin (1).png')
    canvas1.create_image(310, 150, image=img)
    canvas2 = Canvas(frame2, bg='light goldenrod', width=640, height=720)
    canvas2.pack()
    img2 = PhotoImage(file="user.png")
    canvas2.create_image(330, 120, image=img2)
    canvas1.create_text(330, 320, text="Administrator", font=("Times New Roman",30,"bold"))
    canvas2.create_text(330, 260, text="Functionalities provided", font=("Times New Roman",30,"bold"))
    canvas2.create_text(335, 300, text="to the user", font=("Times New Roman",30,"bold"))

    #frame1 buttons
    button1 = Button(frame1, text = "SIGN IN",bg="sienna1",fg="black",activebackground="black",activeforeground="sienna1",font=("Calibri",18,"bold"), command=sign)
    button_canvas1 = canvas1.create_window( 200, 390, anchor = "nw", window = button1,width=250,height=40)
    button2 = Button(frame1, text = "BACK",bg="sienna1",fg="black",activebackground="black",activeforeground="sienna1",font=("Calibri",18,"bold"), command=lambda: usad.destroy())
    button_canvas2 = canvas1.create_window( 200, 460, anchor = "nw", window = button2, width=250,height=40)

    #frame2 buttons
    button3 = Button(frame2, text = "STATIC SCRAPING",bg="indian red",fg="black",activebackground="black",activeforeground="indian red",font=("Calibri",18,"bold"), command=first)
    button_canvas3 = canvas2.create_window( 200, 350, anchor = "nw", window = button3,width=250,height=40)
    button4 = Button(frame2, text = "DYNAMIC SCRAPING",bg="indian red",fg="black",activebackground="black",activeforeground="indian red",font=("Calibri",18,"bold"), command=dscrape)
    button_canvas4 = canvas2.create_window( 200, 420, anchor = "nw", window = button4, width=250,height=40)
    button5 = Button(frame2, text = "REQUEST FOR WEB SCRAPING",bg="indian red",fg="black",activebackground="black",activeforeground="indian red",font=("Calibri",14,"bold"), command=xldb)
    button_canvas5 = canvas2.create_window( 200, 490, anchor = "nw", window = button5, width=250,height=40)
    button6 = Button(frame2, text = "BACK",bg="indian red",fg="black",activebackground="black",activeforeground="indian red",font=("Calibri",18,"bold"), command=lambda: usad.destroy())
    button_canvas6 = canvas2.create_window( 200, 560, anchor = "nw", window = button6, width=250,height=40)

    usad.mainloop()

def sign():
    adm = Toplevel(root)
    adm.title('Login')
    adm.geometry('1280x720')
    adm.configure(bg="#fff")
    os.chdir("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT")
    def signin():
        Admin=user.get()
        Password=code.get()

        os.chdir("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT")
        if Admin=='scott' and Password=='tiger':
            screen=Toplevel(root)
            screen.title("Admin Page")
            screen.geometry('1280x720')
            screen.config(bg='LightSkyBlue1')

            Label(screen,text='Hello Admin!',bg="LightSkyBlue1", font=('Calibri(Body)',50,'bold')).pack()
            Label(screen, text="Here is the information of\nuser requests for web scraping", bg="LightSkyBlue1", font=('Calibri(Body)',40,'bold')).pack(pady=10)
            img1 = PhotoImage(file='3img.png')
            Label(screen,image=img1,bg='LightSkyBlue1').pack(pady=10)

            button100 = Button(screen, text="Open User Details", bg="gold",fg="brown",activebackground="brown",activeforeground="gold",font=("Impact",20,"bold"), command=excel3)
            button100.pack(pady=10)
            button200 = Button(screen, text="Back", bg="gold",fg="brown",activebackground="brown",activeforeground="gold",font=("Impact",20,"bold"), command=lambda: screen.destroy())
            button200.pack(pady=10)

            screen.mainloop()

        elif Admin!='scott' and Password!='tiger':
            messagebox.showerror("Invalid","invalid username and password")

        elif Password!="tiger":
            messagebox.showerror("Invalid","invalid password")

        elif Admin!='scott':
            messagebox.showerror("Invalid","invalid username")
            
    imge = PhotoImage(file='loginn.png')
    Label(adm,image=imge, bg="black").place(x=0,y=0)

    frame=Frame(adm,width=350,height=350,bg="white")
    frame.place(x=480,y=100)

    heading=Label(frame,text='Sign in',fg='#57a1f8',bg='white',font=('Microsoft YaHei UI Light',23,'bold'))
    heading.place(x=100,y=5)

    ############--------------------------------------------------

    def on_enter(e):
        user.delete(0, 'end')

    def on_leave(e):
        name=user.get()
        if name=='':
            user.insert(0,'Admin')
           
    user = Entry(frame,width=25,fg='black',border=0,bg='white',font=('Microsoft YaHei UI Light',11))
    user.place(x=30,y=80)
    user.insert(0,'Admin')
    user.bind('<FocusIn>', on_enter)
    user.bind('<FocusOut>', on_leave)
    Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)

    ############---------------------------------------------------

    def on_enter(e):
        code.delete(0, 'end')

    def on_leave(e):
        name=code.get()
        if name=='':
            code.insert(0,'Password')
    code = Entry(frame,width=25,fg='black',border=0,bg='white',font=('Microsoft YaHei UI Light',11))
    code.place(x=30,y=150)
    code.insert(0,'Password')
    code.bind('<FocusIn>', on_enter)
    code.bind('<FocusOut>', on_leave)

    Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)

    ##############################################################

    Button(frame,width=39,pady=7,text='Sign in',bg='#57a1f8',fg='white',border=0,command=signin).place(x=35,y=204)
    Button(frame,width=39,pady=7,text='Back',bg='#57a1f8',fg='white',border=0,command=lambda: adm.destroy()).place(x=35,y=254)
    if __name__ == "main.py":
        signin()

def excel3():
    file=pathlib.Path('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Full Name"
        sheet['B1']="Contact Number"
        sheet['C1']="Age"
        sheet['D1']="Email"
        sheet['E1']="Url for scraping"
        sheet['F1']="Description"
        sheet['G1']="File Format"
        file.save('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')
    
    # Specify the file path of the Excel workbook to open
    excel_path = 'C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/USER_REQUESTS_FOR_WEB_SCRAPING.xlsx'
        
    # Check if the file exists
    if os.path.isfile(excel_path):
    # Use subprocess to open the file with the default application
        subprocess.Popen(['start', excel_path], shell=True)
    else:
        print('File does not exist')

def dscrape():
    win = Toplevel(root)
    win.title("Scraping and summarizing news articles")
    win.geometry("1280x720")
    win.configure(bg="salmon1")
    def scrape():
        url = urlweb.get()
        article = Article(url)
        authortext.delete("1.0", END)
        dateofpub.delete("1.0", END)
        urlimages.delete("1.0", END)
        summ.delete("1.0", END)
        article.download()
        article.parse()
        article.download('punkt')
        article.nlp()
        authortext.insert(END, f"By {', '.join(article.authors)}\n\n")
        dates = article.publish_date
        if dates is not None:
            datee = dates.strftime("%d/%m/%Y")
            dateofpub.insert(END, datee)
        else:
            dateofpub.insert(END, f"Error: date is None")
        urlimages.insert(END, "\nImages:\n")
        for image in article.images:
            urlimages.insert(END, f"{image}\n")
        summ.insert(END, article.summary)
    
    def clear2():
        urlweb.set('')
        authortext.delete("1.0", END)
        dateofpub.delete("1.0", END)
        urlimages.delete("1.0", END)
        summ.delete("1.0", END)

    #Labels
    headLabel = Label(win, text="Scraping and summarizing news articles", font=("Times New Roman",26,"bold"), bg="salmon1", fg="black")
    headLabel.place(x=360, y=10)
    urlLabel = Label(win, text= "Article URL:", bg="salmon1", fg="black", font=('Calibri',18,'bold'))
    urlLabel.place(x=40, y=80)
    authorLabel = Label(win, text="Author(s):", bg="salmon1", fg="black", font=('Calibri',18,'bold'))
    authorLabel.place(x=40, y=220)
    datepublish = Label(win, text="URLs of\nall images:", bg="salmon1", fg="black", font=('Calibri',18,'bold'))
    datepublish.place(x=40, y=360)
    imageurls = Label(win, text="Date of\npublish:", bg="salmon1", fg="black", font=('Calibri',18,'bold'))
    imageurls.place(x=650, y=220)
    summary = Label(win, text="Summary \nof the\narticle:", bg="salmon1", fg="black", font=('Calibri',18,'bold'))
    summary.place(x=650, y=315)

    #Entry
    urlweb = StringVar()

    #Entrywidget
    urlEntry = Entry(win, textvariable=urlweb, width=90, bd=2, font=18)
    urlEntry.place(x=220, y=85)
    authortext = Text(win)
    authortext.place(x=170, y=220)
    authortext.configure(height=7)
    authortext.configure(width=55)
    urlimages = Text(win)
    urlimages.place(x=170, y=360)
    urlimages.configure(width=55)
    urlimages.configure(height=16)
    dateofpub = Text(win)
    dateofpub.place(x=770, y=220)
    dateofpub.configure(width=55)
    dateofpub.configure(height=3)
    summ = Text(win)
    summ.place(x=770, y=310)
    summ.configure(width=55)
    summ.configure(height=19)

    #Buttons
    buttonn = Button(win, text="Start scraping", font=('Calibri',18,'bold'), bg="misty rose", fg="gray25", command=scrape)
    buttonn.place(x=240, y=140)
    buttonn1 = Button(win, text="Clear", font=('Calibri',18,'bold'), bg="misty rose", fg="gray25", command=clear2)
    buttonn1.place(x=630, y=140)
    buttonn2 = Button(win, text="Exit", font=('Calibri',18,'bold'), bg="misty rose", fg="gray25", command=lambda: win.destroy())
    buttonn2.place(x=950, y=140)

    win.mainloop()

def xldb():
    xld = Toplevel(root)
    xld.title("User Requests Form")
    xld.geometry("1280x720") 
    xld.configure(bg="#326273")

    file=pathlib.Path('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Full Name"
        sheet['B1']="Contact Number"
        sheet['C1']="Age"
        sheet['D1']="Email"
        sheet['E1']="Url for scraping"
        sheet['F1']="Description"
        sheet['G1']="File Format"
        file.save('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')

    def submit():
        name = nameValue.get()
        contact = int(contactValue.get())
        age = int(ageValue.get())
        email = emailValue.get()
        website = websiteValue.get()
        description = descriptionEntry.get(1.0, END)
        files = fileValue.get()

        file=openpyxl.load_workbook('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')
        sheet=file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=name)
        sheet.cell(column=2, row=sheet.max_row, value=contact)
        sheet.cell(column=3, row=sheet.max_row, value=age)
        sheet.cell(column=4, row=sheet.max_row, value=email)
        sheet.cell(column=5, row=sheet.max_row, value=website)
        sheet.cell(column=6, row=sheet.max_row, value=description)
        sheet.cell(column=7, row=sheet.max_row, value=files)

        emptylabel.config(text="Your request will be processed within 2 days\nThank you for approaching us,"+" "+str(name))
        file.save('USER_REQUESTS_FOR_WEB_SCRAPING.xlsx')

    def clear():
        nameValue.set('')
        contactValue.set('')
        websiteValue.set('')
        ageValue.set('')
        emailValue.set('')
        descriptionEntry.delete(1.0, END)
        fileValue.set(" ")

    icon_image = PhotoImage(file="photograph.png")
    xld.iconphoto(False, icon_image)

    #Frame
    Top_frame = Frame(xld, bg="gray6", width=430, height=720, bd=3)
    Top_frame.place(x=850, y=0)
    Top_frame2 = Frame(xld, bg="gray6", width=430, height=220)
    Top_frame2.place(x=850, y=0)

    img100 = PhotoImage(file="ggrp.png")
    Label(Top_frame2, image=img100, bg="gray6").place(x=0, y=0)

    #Heading
    headlabel = Label(xld, text="Please fill out the entry form", font=("Times New Roman",22,"bold"), bg="#326273", fg="#fff")
    headlabel.place(x=20, y=20)

    #Label
    nameLabel = Label (xld, text='Name' , font=23, bg="#326273", fg="#fff")
    nameLabel.place(x=50,y=100)
    contactLabel = Label(xld, text='Contact No. ', font=23, bg="#326273", fg="#Fff")
    contactLabel.place(x=50,y=160)
    ageLabel = Label(xld, text='Age', font=23, bg="#326273", fg="#fff")
    ageLabel.place(x=50,y=220)
    emailLabel = Label(xld ,text=' Email', font=23, bg="#326273", fg="#fff")
    emailLabel.place(x=290,y=220)
    websiteLabel = Label(xld, text="URL of website\nto be scraped", font=23, bg="#326273", fg="#fff")
    websiteLabel.place(x=50, y=280)
    descriptionLabel = Label(xld ,text='Provide description\nabout your request' , font=23, bg="#326273", fg="#fff")
    descriptionLabel.place(x=50,y=340)
    fileformatLabel = Label(xld, text="Choose the file\nformat in which\nthe results to\n be stored", font=23, bg="#326273", fg="#fff")
    fileformatLabel.place(x=50, y=500)

    #Entry
    nameValue = StringVar()
    contactValue = StringVar()
    websiteValue = StringVar()
    ageValue = StringVar()
    emailValue = StringVar()
    fileValue=StringVar(value=" ")

    #Entrywidget
    nameEntry = Entry(xld, textvariable=nameValue, width=45, bd=2, font=20)
    nameEntry.place(x=200, y=100)
    contactEntry = Entry(xld, textvariable=contactValue, width=45, bd=2, font=20)
    contactEntry.place(x=200, y=160)
    ageEntry = Entry(xld, textvariable=ageValue, width=6, bd=2, font=20)
    ageEntry.place(x=200, y=220)
    emailEntry = Entry(xld, textvariable=emailValue, width=25, bd=2, font=20)
    emailEntry.place(x=380, y=220)
    websiteEntry = Entry(xld, textvariable=websiteValue, width=45, bd=2, font=20)
    websiteEntry.place(x=200, y=280)
    descriptionEntry = Text(xld, width=51, height=8, bd=2)
    descriptionEntry.place(x=200, y=340)

    #Buttons
    button1 = Button(Top_frame, text="Submit", bg="gold", fg="brown", width=12, height=2, font=("Arial",14,"bold"), command=submit)
    button1.place(x=150, y=300)
    button2 = Button(Top_frame, text="Clear", bg="gold", fg="brown", width=12, height=2, font=("Arial",14,"bold"), command=clear)
    button2.place(x=150, y=400)
    button3 = Button(Top_frame, text="Exit", bg="gold", fg="brown", width=12, height=2, font=("Arial",14,"bold"), command=lambda: xld.destroy())
    button3.place(x=150, y=500)

    #Radiobutton
    radio1 = Radiobutton(xld, text="Excel", value="Excel", variable=fileValue)
    radio1.place(x=200, y=500)
    radio2 = Radiobutton(xld, text="CSV", value="CSV", variable=fileValue)
    radio2.place(x=300, y=500)

    #emptylabel
    emptylabel = Label(xld, bg="#326273", fg="orange", font=("Times New Roman",18,"bold"))
    emptylabel.place(x=190, y=560)

    xld.mainloop()

def open_excel():
    # Specify the file path of the Excel workbook to open
    excel_path = 'C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/Top_Programming_Languages.xlsx'
        
    # Check if the file exists
    if os.path.isfile(excel_path):
    # Use subprocess to open the file with the default application
        subprocess.Popen(['start', excel_path], shell=True)
    else:
        print('File does not exist')

def open_excel2():
    # Specify the file path of the Excel workbook to open
    excel_path = 'C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/IMDB_Movie_Ratings.xlsx'
    
    # Check if the file exists
    if os.path.isfile(excel_path):
        # Use subprocess to open the file with the default application
        subprocess.Popen(['start', excel_path], shell=True)
    else:
        print('File does not exist')

def first():
    window1 = Toplevel(root)
    # Create two frames, one for each half of the window
    frame1 = Frame(window1)
    frame1.grid(row=0, column=0, sticky='nsew')
    frame2 = Frame(window1)
    frame2.grid(row=0, column=1, sticky='nsew')

    # Create two canvases, one in each frame
    canvas1 = Canvas(frame1, bg='orange', width=640, height=720)
    canvas1.pack()
    img = PhotoImage(file='99.png')
    canvas1.create_image(310, 100, image=img)
    canvas2 = Canvas(frame2, bg='lightblue', width=640, height=720)
    canvas2.pack()
    img2 = PhotoImage(file="imdb.png")
    canvas2.create_image(310, 100, image=img2)
    canvas1.create_text(320, 250, text="Let's scrape 99firms website", font=("Times New Roman",30,"bold"))
    canvas2.create_text(320, 250, text="Let's scrape IMDB website", font=("Times New Roman",30,"bold"))

    #frame1 buttons
    button1 = Button(frame1, text = "START SCRAPING",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"), command = second)
    button_canvas1 = canvas1.create_window( 200, 350, anchor = "nw", window = button1,width=250,height=40)
    button2 = Button(frame1, text = "BACK",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"), command=lambda: window1.destroy())
    button_canvas2 = canvas1.create_window( 200, 450, anchor = "nw", window = button2, width=250,height=40)
    #frame2 buttons
    button3 = Button(frame2, text = "START SCRAPING",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"), command= third)
    button_canvas3 = canvas2.create_window( 200, 350, anchor = "nw", window = button3,width=250,height=40)
    button4 = Button(frame2, text = "BACK",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"), command=lambda: window1.destroy())
    button_canvas4 = canvas2.create_window( 200, 450, anchor = "nw", window = button4, width=250,height=40)

    window1.mainloop()

def second():
    window2 = Toplevel(root)
    window2.title("Scraping 99frims website")
    window2.geometry("1280x720")
    result = some_function()

    #Define image
    bg = ImageTk.PhotoImage(file="img2.png")
    bg2 = ImageTk.PhotoImage(file="img3.png")
    bg3 = ImageTk.PhotoImage(file="img4.png")
    bg4 = ImageTk.PhotoImage(file="img5.png")
    bg5 = ImageTk.PhotoImage(file="img6.png")
    bg6 = ImageTk.PhotoImage(file="img7.png")

    # Create a canvas
    canvas3 = Canvas(window2, width=1280, height=720)
    canvas3.pack(fill="both", expand=True)
    # Set image in canvas
    canvas3.create_image(0,0, image=bg, anchor="nw")
    canvas3.create_image(640,150, image=bg2)
    canvas3.create_image(640,300, image=bg3)
    canvas3.create_image(320, 430, image=bg4)
    canvas3.create_image(640, 430, image=bg5)
    canvas3.create_image(960, 430, image=bg6)

    # Add a label
    canvas3.create_text(645, 35, text = "HERE ARE THE TOP PROGRAMMING LANGUAGES",font=("Times New Roman",36,"bold"),fill="lemon chiffon")
    canvas3.create_text(650, 90,text=result,fill="lemon chiffon",font=("Times", 28, "bold"))

    #button
    button = Button(window2,text = "Open Excel Workbook",bg="gold",fg="brown",activebackground="brown",activeforeground="gold",font=("Impact",20,"bold"),relief=SOLID, command=open_excel)
    button_canvas = canvas3.create_window( 640, 210, window = button,width=300,height=40)
    button = Button(window2,text = "BAR CHART",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=barch)
    button_canvas = canvas3.create_window( 200, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window2,text = "PIE CHART",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=piech)
    button_canvas = canvas3.create_window( 520, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window2,text = "SWARM PLOT",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=swarmch)
    button_canvas = canvas3.create_window( 840, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window2,text = "BACK",bg="green2",fg="black",font=("Impact",20,"bold"),relief=SOLID, command=lambda: window2.destroy())
    button_canvas = canvas3.create_window( 640, 580, window = button,width=240,height=40)
    window2.mainloop()

def third():
    window3 = Toplevel(root)
    window3.title("Scraping IMDB website")
    window3.geometry("1280x720")
    result2 = some_function2()

    #Define image
    bg = ImageTk.PhotoImage(file="img8.png")
    bg2 = ImageTk.PhotoImage(file="img3.png")
    bg3 = ImageTk.PhotoImage(file="img4.png")
    bg4 = ImageTk.PhotoImage(file="img9.png")
    bg5 = ImageTk.PhotoImage(file="img10.png")
    bg6 = ImageTk.PhotoImage(file="img11.png")

    # Create a canvas
    canvas4 = Canvas(window3, width=1280, height=720)
    canvas4.pack(fill="both", expand=True)
    # Set image in canvas
    canvas4.create_image(0,0, image=bg, anchor="nw")
    canvas4.create_image(640,150, image=bg2)
    canvas4.create_image(640,300, image=bg3)
    canvas4.create_image(320, 430, image=bg4)
    canvas4.create_image(640, 430, image=bg5)
    canvas4.create_image(960, 430, image=bg6)

    # Add a label
    canvas4.create_text(645, 35, text = "HERE ARE THE TOP 250 MOVIES RATED BY REGULAR IMDB VOTERS",font=("Times New Roman",25,"bold"),fill="lemon chiffon")
    canvas4.create_text(650, 90,text=result2,fill="lemon chiffon",font=("Times", 28, "bold"))

    #button
    button = Button(window3,text = "Open Excel Workbook",bg="gold",fg="brown",activebackground="brown",activeforeground="gold",font=("Impact",20,"bold"),relief=SOLID, command=open_excel2)
    button_canvas = canvas4.create_window( 640, 210, window = button,width=300,height=40)
    button = Button(window3,text = "SCATTER PLOT",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=scatterpl)
    button_canvas = canvas4.create_window( 200, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window3,text = "HISTOGRAM",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=histo)
    button_canvas = canvas4.create_window( 520, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window3,text = "REGRESSION PLOT",bg="brown",fg="gold",activebackground="gold",activeforeground="black",font=("Impact",20,"bold"), command=regressionpl)
    button_canvas = canvas4.create_window( 840, 500, anchor = "nw", window = button,width=240,height=40)
    button = Button(window3,text = "BACK",bg="green2",fg="black",font=("Impact",20,"bold"),relief=SOLID, command=lambda: window3.destroy())
    button_canvas = canvas4.create_window( 640, 580, window = button,width=240,height=40)
    #window3.overrideredirect(True)

    window3.mainloop()

def some_function():
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = "Top-Programming-Languages"
    sheet.append(['Programming_Language','Market_Share_in_Percentage'])

    source = requests.get("https://99firms.com/blog/most-popular-programming-languages/#gref")
    soup = BeautifulSoup(source.text, 'html.parser')
    titles = soup.find_all('h3')[0:10]
    titl = []
    for m in titles:
        titl.append(m.string)
    shares = soup.find_all('p')[1:11]
    shar = []
    for n in shares:
        shar.append(float(n.string.split(' ')[3].replace("%", "")))
    contents1 = soup.find('h2')
    contents = soup.find_all('p')[1:11]
    i=0
    while(i<10):
        sheet.append([titl[i],float(shar[i])])
        i+=1
    excel.save('Top_Programming_Languages.xlsx')
    return f"The required data is scraped and saved in an excel workbook in the system!"

def some_function2():
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = "Top_Rated_Movies"
    sheet.append(['Movie_Rank','Movie_Name','Year_of_Release','IMDB_Rating'])

    source = requests.get("https://www.imdb.com/chart/top/")
    soup = BeautifulSoup(source.text, 'html.parser')
    movies = soup.find('tbody', class_="lister-list").find_all('tr')
    ran = []
    nam = []
    yea = []
    ratin = []
    for movie in movies:
        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
        name = movie.find('td', class_="titleColumn").a.text
        year = movie.find('td', class_="titleColumn").span.text.strip('()')
        rating = movie.find('td', class_="ratingColumn imdbRating").strong.text
        ran.append(int(rank))
        nam.append(name)
        yea.append(int(year))
        ratin.append(float(rating))
    i=0
    while(i<250):
        sheet.append([ran[i],nam[i],yea[i],ratin[i]])
        i+=1
    excel.save('IMDB_Movie_Ratings.xlsx')
    return f"The required data is scraped and saved in an excel workbook in the system"

def barch():
    import pandas as pd
    import matplotlib.pyplot as plt

    # read the excel file and store it in a DataFrame
    df = pd.read_excel("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/Top_Programming_Languages.xlsx")
    # create a bar chart of the data
    ax = df.plot(kind='bar', x='Programming_Language', y='Market_Share_in_Percentage')

    # add a title and labels
    plt.title("Top_Programming_Languages (Bar Chart)")
    plt.xlabel("Programming_Language")
    plt.ylabel("Market_Share_in_Percentage")
    # add the data labels to the bars
    for i in ax.containers:
        ax.bar_label(i)

    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    # show the chart
    plt.show()

def piech():
    import matplotlib.pyplot as plt
    import pandas as pd

    # read the excel file and store it in a DataFrame
    df = pd.read_excel("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/Top_Programming_Languages.xlsx")
    # data to plot
    labels = df['Programming_Language']
    data = df['Market_Share_in_Percentage']
    # Pie chart, where the slices will be ordered and plotted counter-clockwise:
    fig1, ax1 = plt.subplots(figsize=(12,12))
    ax1.pie(data, labels=labels, autopct='%.1f%%',startangle=90)

    # change the font size of the labels
    for text in ax1.texts:
        text.set_fontsize(6)
    # Equal aspect ratio ensures that pie is drawn as a circle
    ax1.axis('equal')
    #add a title to the chart
    plt.title("Top_Programming_Languages (Pie Chart)")

    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    plt.show()

def swarmch():
    import seaborn as sns
    import matplotlib.pyplot as plt
    import os

    os.chdir("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT")
    # Load the data into a Pandas DataFrame
    import pandas as pd
    data = pd.read_excel("Top_Programming_Languages.xlsx")
    # Plot the Swarm plot using Seaborn
    sns.swarmplot(x='Programming_Language', y='Market_Share_in_Percentage', data=data)

    # Add data labels to the data points
    for line in range(0, data.shape[0]):
        plt.text(x = data['Programming_Language'][line], y = data['Market_Share_in_Percentage'][line], s = data['Market_Share_in_Percentage'][line], size = 8)
    #add a title to the plot
    plt.title("Top_Programming_Languages (Swarm Plot)")

    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    # Show the plot
    plt.show()

def scatterpl():
    import pandas as pd
    import matplotlib.pyplot as plt

    # Load the data from a local file into a Pandas DataFrame
    data = pd.read_excel("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/IMDB_Movie_Ratings.xlsx")
    # Create the scatterplot
    plt.scatter(data['Year_of_Release'], data['IMDB_Rating'])

    #add a title to the plot
    plt.title("IMDB_Movie_Ratings (Scatter Plot)")
    # Add labels to the axes
    plt.xlabel("Year of Release")
    plt.ylabel("IMDB Rating")

    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    plt.show()

def histo():
    import matplotlib.pyplot as plt
    import pandas as pd
    import seaborn as sns

    # Load the dataset
    data = pd.read_excel("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/IMDB_Movie_Ratings.xlsx")
    # Plot the histogram of the 'Market_Share_in_Percentage' column
    plt.hist(data['Year_of_Release'], bins=50, edgecolor='black', linewidth=1.2)

    #add a title to the plot
    plt.title("IMDB_Movie_Ratings (Histogram)")
    #x-axis label
    plt.xlabel("Year of Release")

    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    plt.show()

def regressionpl():
    import pandas as pd
    import seaborn as sns
    import matplotlib.pyplot as plt

    # Load the data from a local file into a Pandas DataFrame
    data = pd.read_excel("C:/Users/JASHWANTH/Downloads/JASHWANTH_LOYOLA_MAJORPROJECT/IMDB_Movie_Ratings.xlsx")
    # Create the regression plot
    sns.regplot(x='Year_of_Release', y='IMDB_Rating', data=data)

    #add a title to the chart
    plt.title("IMDB_Movie_Ratings (Regression Plot)")
    # Set the title for the plot window
    fig_manager = plt.get_current_fig_manager()
    fig_manager.window.state('zoomed')
    # Show the plot
    plt.show()

root = Tk()
root.title("Web Scraping using Python")
root.geometry("1280x720")

# Define image
bg = ImageTk.PhotoImage(file="img.png")

# Create a canvas
my_canvas = Canvas(root, width=1280, height=720)
my_canvas.pack(fill="both", expand=True)

# Set image in canvas
my_canvas.create_image(0,0, image=bg, anchor="nw")

# Add a label
my_canvas.create_text(645, 40, text = "Web Scraping using Python",font=("Times New Roman",40,"bold underline"),fill="black")
my_canvas.create_text(650, 90,text="Programmed by Jashwanth and Rakshith",fill="black",font=("Times", 24, "bold"))

#button
button = Button(root,text = "LET'S GET STARTED!",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"),relief=SOLID, command=usadmin)
button_canvas = my_canvas.create_window( 300, 580, anchor = "nw", window = button,width=250,height=40)
button = Button(root,text = "EXIT THE APPLICATION",bg="red",fg="white",activebackground="green2",activeforeground="black",font=("Impact",20,"bold"),relief=SOLID, command=lambda: root.destroy())
button_canvas = my_canvas.create_window( 750, 580, anchor = "nw", window = button,width=270,height=40)

#to run event loop
root.mainloop()